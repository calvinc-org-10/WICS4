from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel, WINDOWS_EPOCH



def test_read():
    fName = r"D:\tmp0\Counts Entered 2026-07-20to.xlsx"
    wb = load_workbook(filename=fName, read_only=True, data_only=True)
    CountSprshtDateEpoch = wb.epoch
    if 'Counts' in wb:
        ws = wb['Counts']
    else:
        ws = None

    assert ws is not None, "Spreadsheet does not have a 'Counts' worksheet"
    SprshtcolmnNames = ws[1]
    SprshtREQUIREDFLDS = ['Material','CountDate','Counter','LOCATION']
        # LocationOnly/CTD_QTY_Expr handled separately since at least one must be present and both can be
    SprshtcolmnMap = {}
    Sprsht_SSName_TableName_map = {
            'CountDate': 'CountDate',
            'Counter': 'Counter',
            'LOCATION': 'LOCATION',
            'org_id': 'org_id',
            'Material': 'Material',
            'LocationOnly': 'LocationOnly',
            'CTD_QTY_Expr': 'CTD_QTY_Expr',
            'Typ Cntner Qty': 'TypicalContainerQty',
            'Typ Plt Qty': 'TypicalPalletQty',
            'Notes': 'Notes',
            'PKGID_Desc': 'PKGID_Desc',
            'TAGQTY': 'TAGQTY',
            'Poss Not Rcvd': 'FLAG_PossiblyNotRecieved',
            'Mvmt Dur Ct': 'FLAG_MovementDuringCount',
            'WICSignore': 'FLAG_WICSIgnore',
            }
    for col in SprshtcolmnNames:
        if col.value in Sprsht_SSName_TableName_map:
            colkey = Sprsht_SSName_TableName_map[str(col.value)]
            assert col.column is not None, f"Column {colkey} has no column number"
            SprshtcolmnMap[colkey] = col.column - 1
            # endif previously mapped
        #endif col.value in SAP_SSName_TableName_map
    #endfor col in SAPcolmnNames

    HeaderGood = all([(reqFld in SprshtcolmnMap) for reqFld in SprshtREQUIREDFLDS])

    SprshtRowNum=1
    
    for row in ws.iter_rows(min_row=SprshtRowNum+1, values_only=True):
        SprshtRowNum += 1

    return f"Read {SprshtRowNum} rows from spreadsheet {fName}, header good: {HeaderGood}"

