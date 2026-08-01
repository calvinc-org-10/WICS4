import uuid, os, re as regex, ast
from enum import Enum

from flask import (
    current_app,
    session, 
    redirect, url_for,
    request, make_response, jsonify,
    )
from flask_login import login_required
from flask_wtf import FlaskForm

from sqlalchemy import text, inspect
from sqlalchemy.sql import select

from openpyxl import load_workbook

# from async_tasks import huey

from calvincTools.utils import (
    checkTemplate_and_render,
    ExcelWorkbook_fileext,
    )

from database import (app_db,)
from models import (
    tmpMaterialListUpdate, SAPPlants_org,
    async_comm
    )
    

####################################################################################
####################################################################################
####################################################################################

##### the suite of procs to support fnUpdateMatlListfromSAP

class FatalUploadError(Exception):
    pass

def proc_MatlListSAPSprsheet_00InitUMLasync_comm(reqid, UpdateExistFldList, rmvMissingMaterial=False):
    # these first calls should create the async_comm record with pk=reqid.  All subsequent calls will update that same record until we delete it in the cleanup proc at the end.
    acomm = async_comm.set_async_comm_state(
        reqid,
        statecode = 'rdng-sprsht-init',
        statetext = 'Initializing ...',
        )   
    async_comm.set_async_comm_state(
        f'{reqid}-UpdExstFldList',
        statecode = 'UpdateExistFldList',
        statetext = f'{UpdateExistFldList}',
        )
    async_comm.set_async_comm_state(
        f'{reqid}-RmvMissingMatl',
        statecode = 'RmvMissingMatl',
        statetext = f'{rmvMissingMaterial}',
        )

def proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet(reqid, uselocalCopy=False):
    acomm = async_comm.set_async_comm_state(
        reqid,
        statecode = 'uploading-sprsht',
        statetext = 'Uploading Spreadsheet',
        )

    SAPFile = request.files.get('SAPFile')
    if SAPFile is None:
        acomm = async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = 'No file uploaded. Please upload an Excel spreadsheet and try again.',
            result = 'FAIL - no file uploaded',
            )
        return
    svdir = current_app.config.get('SAP_FILELOC', os.getcwd()) if not uselocalCopy else ''
    os.makedirs(svdir, exist_ok=True)    
    fName = svdir+"tmpMatlList"+str(reqid)+ExcelWorkbook_fileext
    SAPFile.save(fName)

    return fName

def proc_MatlListSAPSprsheet_00ResolveLocalSpreadsheetPath(reqid):
    local_path = (request.form.get('SAPFileServerPath', '') or '').strip()

    if not local_path:
        statetext = 'No server file path provided for local copy mode.'
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = statetext,
            result = 'FAIL - missing local path',
            )
        raise FatalUploadError(statetext)

    if not os.path.isabs(local_path):
        statetext = f'Local spreadsheet path must be absolute: {local_path}'
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = statetext,
            result = 'FAIL - non-absolute local path',
            )
        raise FatalUploadError(statetext)

    if not os.path.exists(local_path):
        statetext = f'Local spreadsheet file not found: {local_path}'
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = statetext,
            result = 'FAIL - local path not found',
            )
        raise FatalUploadError(statetext)

    if os.path.isdir(local_path):
        statetext = f'Local spreadsheet path points to a directory, not a file: {local_path}'
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = statetext,
            result = 'FAIL - local path is a directory',
            )
        raise FatalUploadError(statetext)

    _, ext = os.path.splitext(local_path)
    if ext.lower() != ExcelWorkbook_fileext.lower():
        statetext = f'Local spreadsheet must be an {ExcelWorkbook_fileext} file: {local_path}'
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = statetext,
            result = 'FAIL - invalid local file extension',
            )
        raise FatalUploadError(statetext)

    return local_path

# @huey.context_task(thisapp.app_context())
def proc_MatlListSAPSprsheet_01ReadSpreadsheet(reqid, fName, cleanup_file=True):
    acomm = async_comm.set_async_comm_state(
        reqid,
        statecode = 'rdng-sprsht',
        statetext = 'Reading Spreadsheet',
        )

    if len(fName)<1 or not os.path.exists(fName):
        statetext = f'Spreadsheet file {fName} not found. Please try again.'
        acomm = async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = statetext,
            result = 'FAIL - file not found',
            )
        raise FatalUploadError(statetext)
    
    app_db.session.query(tmpMaterialListUpdate).delete(synchronize_session=False)
    app_db.session.commit()
    
    wb = load_workbook(filename=fName, read_only=True)
    ws = wb.active
    assert ws is not None, "Error: Spreadsheet appears to be blank. Please fix this and try again."
    SAPcolmnNames = ws[1]
    SAPcol:dict[str, int|None] = {'Plant':None,'Material': None}
    SAP_SSName_TableName_map = {
            'Material': 'Material',
            'Material description': 'Description',
            'Plant': 'Plant', 'Plnt': 'Plant',
            'Material type': 'SAPMaterialType',  'MTyp': 'SAPMaterialType',
            'Material Group': 'SAPMaterialGroup', 'Matl Group': 'SAPMaterialGroup',
            'Manufact.': 'SAPManuf', 
            'MPN': 'SAPMPN', 
            'ABC': 'SAPABC', 
            'Price': 'Price', 'Standard price': 'Price',
            'Price unit': 'PriceUnit', 'per': 'PriceUnit',
            'Currency':'Currency',
            }
    for col in SAPcolmnNames:
        assert col.value is not None, f"Error: Blank column header found in spreadsheet at column {col.column}. Please fix this and try again."
        assert col.column is not None, f"Error: Column with blank header has no column number. This shouldn't happen. Please check the spreadsheet and try again."
        if col.value in SAP_SSName_TableName_map:
            colval = str(col.value)
            SAPcol[SAP_SSName_TableName_map[colval]] = col.column - 1
    if (SAPcol['Material'] == None or SAPcol['Plant'] == None):
        statetext = 'SAP Spreadsheet has bad header row. Plant and/or Material is missing.  See Calvin to fix this.'
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = statetext,
            result = 'FAIL - bad spreadsheet',
            )

        wb.close()
        if cleanup_file and os.path.exists(fName):
            os.remove(fName)
        raise FatalUploadError(statetext)

    numrows = ws.max_row
    nRows = 0
    reportEveryNRows = min(100, max(1, numrows//10))
    for row in ws.iter_rows(min_row=2, values_only=True):
        nRows += 1
        if nRows % reportEveryNRows == 0:
            async_comm.set_async_comm_state(
                reqid,
                statecode = 'rdng-sprsht',
                statetext = f'Reading Spreadsheet ... record {nRows} of {numrows}<br><progress max="{numrows}" value="{nRows}"></progress>',
                )

        if row[SAPcol['Material']]==None: MatNum = ''
        else: MatNum = row[SAPcol['Material']]
        validTmpRec = False
        ## create a blank tmpMaterialListUpdate record,
        newrec = tmpMaterialListUpdate()
        if regex.match(".*[\n\t\xA0].*",str(MatNum)):
            validTmpRec = True
            ## refuse to work with special chars embedded in the MatNum
            newrec.recStatus = 'err-MatlNum'
            newrec.errmsg = f'error: {MatNum!a} is an unusable part number. It contains invalid characters and cannot be added to WICS'
        elif len(str(MatNum)):
            validTmpRec = True
            plant_col = SAPcol['Plant']
            if plant_col is not None:
                org_row = SAPPlants_org.query.filter_by(SAPPlant=row[plant_col]).first()
                if org_row is not None:
                    newrec.org = org_row.org
        # endif invalid Material
        if validTmpRec:
            ## populate by looping through SAPcol,
            ## then save
            for dbColName, ssColNum in SAPcol.items():
                assert ssColNum is not None, f"Error: Column {dbColName} has no column number. This shouldn't happen. Please check the spreadsheet and try again."
                setattr(newrec,dbColName,row[ssColNum])
            
            app_db.session.add(newrec)
            app_db.session.commit()
    # endfor

    wb.close()
    if cleanup_file and os.path.exists(fName):
        os.remove(fName)

    # report done and move to next step
    statecode = getattr(async_comm.get_async_comm_state(reqid), 'statecode', 'fatalerr')                 # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to fatalerr if we can't get the statecode
    statetext = getattr(async_comm.get_async_comm_state(reqid), 'statetext', f'No state for {reqid}')    # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to fatalerr if we can't get the statecode
    if statecode != 'fatalerr':
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'done-rdng-sprsht',
            statetext = f'Finished Reading Spreadsheet',
            )
        return reqid
    else:
        statetext = f'Error: Something went wrong while reading the spreadsheet. Please check the spreadsheet and try again. Details: {statetext}'
        # async_comm.set_async_comm_state(
        #     reqid,
        #     statecode = 'fatalerr',
        #     statetext = statetext,
        #     result = 'FAIL - error reading spreadsheet',
        #     )
        raise FatalUploadError(statetext)
# proc_MatlListSAPSprsheet_01ReadSpreadsheet

# @huey.context_task(thisapp.app_context())
def proc_MatlListSAPSprsheet_02_identifyexistingMaterial(reqid):
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'get-matl-link',
        statetext = f'Finding SAP MM60 Materials already in WICS Material List',
        )
    UpdMaterialLinkSQL = 'UPDATE WICS_tmpmateriallistupdate, (select id, org_id, Material from WICS_materiallist) as MasterMaterials'
    UpdMaterialLinkSQL += ' set WICS_tmpmateriallistupdate.MaterialLink_id = MasterMaterials.id, '
    UpdMaterialLinkSQL += "     WICS_tmpmateriallistupdate.recStatus = 'FOUND' "
    UpdMaterialLinkSQL += ' where WICS_tmpmateriallistupdate.org_id = MasterMaterials.org_id '
    UpdMaterialLinkSQL += '   and WICS_tmpmateriallistupdate.Material = MasterMaterials.Material '
    app_db.session.execute(text(UpdMaterialLinkSQL))
    app_db.session.commit()

    rmvMissingMaterial = getattr(async_comm.get_async_comm_state(f"{reqid}-RmvMissingMatl"), 'statetext', 'False')
    if rmvMissingMaterial in ['True', True]:
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'id-del-matl',
            statetext = f'Identifying WICS Materials no longer in SAP MM60 Materials',
            )
        MustKeepMatlsSelCond = ''
        MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
        MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT tmucopy.MaterialLink_id AS Material_id FROM WICS_tmpmateriallistupdate tmucopy WHERE tmucopy.MaterialLink_id IS NOT NULL)'
        MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
        MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_actualcounts)'
        MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
        MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_countschedule)'
        MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
        MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_sap_sohrecs)'

        DeleteMatlsSelectSQL = "INSERT INTO WICS_tmpmateriallistupdate (recStatus, delMaterialLink, MaterialLink_id, org_id, Material, Description, Plant "
        DeleteMatlsSelectSQL += ", SAPMaterialType, SAPMaterialGroup, Currency  ) "    # these can go once I set null=True on these fields
        DeleteMatlsSelectSQL += " SELECT  concat('DEL ',FORMAT(id,0)), id, NULL, org_id, Material, Description, Plant "
        DeleteMatlsSelectSQL += ", SAPMaterialType, SAPMaterialGroup, Currency  "    # these can go once I set null=True on these fields
        DeleteMatlsSelectSQL += " FROM WICS_materiallist"
        DeleteMatlsSelectSQL += f" WHERE ({MustKeepMatlsSelCond})"
        app_db.session.execute(text(DeleteMatlsSelectSQL))
        app_db.session.commit()
    # end if rmvMissingMaterial in ('True', True)

    async_comm.set_async_comm_state(
        reqid,
        statecode = 'id-add-matl',
        statetext = f'Identifying SAP MM60 Materials new to WICS',
        )
    MarkAddMatlsSelectSQL = "UPDATE WICS_tmpmateriallistupdate"
    MarkAddMatlsSelectSQL += " SET recStatus = 'ADD'"
    MarkAddMatlsSelectSQL += " WHERE (MaterialLink_id IS NULL) AND (recStatus is NULL)"
    app_db.session.execute(text(MarkAddMatlsSelectSQL))
    app_db.session.commit()

    # report done and move to next step
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'get-matl-link-done',
        statetext = f'Finished linking SAP MM60 list to existing WICS Materials',
        )
    
    return reqid
# proc_MatlListSAPSprsheet_02_identifyexistingMaterial

# @huey.context_task(thisapp.app_context())
def proc_MatlListSAPSprsheet_03_UpdateExistingRecs(reqid):
    def setstate_MatlListSAPSprsheet_03_UpdateExistingRecs(fldName):
        acomm = async_comm.set_async_comm_state(
            reqid,
            statecode = 'upd-existing-recs',
            statetext = f'Updating _{fldName}_ Field in Existing Records',
            )

    setstate_MatlListSAPSprsheet_03_UpdateExistingRecs('')

    # (Form Name, db fld Name, zero/blank value)
    FormTodbFld_map = [
        ('Description','Description','""'),
        ('SAPMatlType','SAPMaterialType','""'),
        ('SAPMatlGroup','SAPMaterialGroup','""'),
        ('SAPManuf','SAPManuf','""'),
        ('SAPMPN','SAPMPN','""'),
        ('SAPABC','SAPABC','""'),
        ('SAPPrice','Price',0),
        ('SAPPrice','PriceUnit',0),
        ('SAPPrice','Currency','""'),
    ]

    UpdateExistFldList_str = getattr(async_comm.get_async_comm_state(f"{reqid}-UpdExstFldList"), 'statetext', '[]')    # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to '' if we can't get the statetext
    UpdateExistFldList = ast.literal_eval(UpdateExistFldList_str)

    if UpdateExistFldList:
        for formName, dbName, zeroVal in FormTodbFld_map:
            if formName in UpdateExistFldList:
                setstate_MatlListSAPSprsheet_03_UpdateExistingRecs(dbName)
                # UPDATE this field
                UpdSQLSetStmt = f"MatlList.{dbName}=tmpMatl.{dbName}"
                UpdSQLWhereStmt = f"(IFNULL(tmpMatl.{dbName},{zeroVal}) != {zeroVal} AND IFNULL(MatlList.{dbName},{zeroVal})!=IFNULL(tmpMatl.{dbName},{zeroVal}))"

                UpdSQLStmt = "UPDATE WICS_materiallist AS MatlList, WICS_tmpmateriallistupdate AS tmpMatl"
                UpdSQLStmt += f" SET {UpdSQLSetStmt}"
                UpdSQLStmt += f" WHERE (tmpMatl.MaterialLink_id=MatlList.id) AND {UpdSQLWhereStmt}"
                app_db.session.execute(text(UpdSQLStmt))
                app_db.session.commit()
            #endif formName in UpdateExistFldList
        #endfor
    # endif UpdateExistFldList not empty

    # report done and move to next step
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'upd-existing-recs-done',
        statetext = f'Finished Updating Existing Records to MM60 values',
        )
    return reqid
# proc_MatlListSAPSprsheet_03_UpdateExistingRecs

# @huey.context_task(thisapp.app_context())
def proc_MatlListSAPSprsheet_04_Remove(reqid):
# temporarily skipped ...

    doRmv_str = getattr(async_comm.get_async_comm_state(f"{reqid}-RmvMissingMatl"), 'statetext', 'False')
    doRmv = doRmv_str in [True, 'True'] # ast.literal_eval(doRmv_str)

    if not doRmv:
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'del-matl-skip',
            statetext = f'Not Removing WICS Materials no longer in SAP MM60 Materials',
            )
        proc_MatlListSAPSprsheet_04_Add(reqid)
        return
    else:
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'del-matl',
            statetext = f'Removing WICS Materials no longer in SAP MM60 Materials',
            )

        # do the Removals
        ## DeleteMatlsDoitSQL = "DELETE FROM WICS_materiallist"
        ## DeleteMatlsDoitSQL += f" WHERE ({MustKeepMatlsDelCond})"
        DeleteMatlsDoitSQL = 'DELETE MATL'
        DeleteMatlsDoitSQL += ' FROM WICS_materiallist AS MATL INNER JOIN WICS_tmpmateriallistupdate AS TMP'
        DeleteMatlsDoitSQL += '    ON MATL.id = TMP.delMaterialLink'
        DeleteMatlsDoitSQL += ' WHERE TMP.recStatus like "DEL%"'
        app_db.session.execute(text(DeleteMatlsDoitSQL))
        app_db.session.commit()
    # endif doRmv

    # report done and move to next step    
    mandatorytaskdonekey = f'MatlX{reqid}'
    statecodeVal = ".03D."
    existingstatecode = ''
    if async_comm.async_comm_exists(mandatorytaskdonekey): 
        existingstatecode = getattr(async_comm.get_async_comm_state(mandatorytaskdonekey), 'statecode', '')    # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to '' if we can't get the statecode
    MatlXval = existingstatecode + statecodeVal
    async_comm.set_async_comm_state(
        mandatorytaskdonekey,
        statecode = MatlXval,
        statetext = '',
        )
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'done-del-matl',
        statetext = 'Finished Removing' if doRmv else 'Skipped Removal of' + 'WICS Materials no longer in SAP MM60 Materials',
        )
    return reqid
# proc_MatlListSAPSprsheet_04_Remove
# @huey.context_task(thisapp.app_context())
def proc_MatlListSAPSprsheet_04_Add(reqid):
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'add-matl',
        statetext = f'Adding SAP MM60 Materials new to WICS',
        )
    # phase out the UNKNOWN type; just leave WhsePartType_id blank for new records until we can get a real type assigned in WICS
    # UnknownTypeID = WhsePartTypes.objects.using(dbToUse).get(WhsePartType=WICS.globals._PartTypeName_UNKNOWN)

    # do the adds
    # one day django will implement insert ... select.  Until then ...
    # come back to this one day and rewrite it to use the ORM instead of raw SQL;
    AddMatlsSelectSQL = "SELECT"
    # AddMatlsSelectSQL += " org_id, Material, Description, Plant, " + str(UnknownTypeID.pk) + " AS PartType_id,"
    AddMatlsSelectSQL += " org_id, Material, Description, Plant,"
    AddMatlsSelectSQL += " SAPMaterialType, SAPMaterialGroup, Price, PriceUnit, Currency,"
    AddMatlsSelectSQL += " '' AS TypicalContainerQty, '' AS TypicalPalletQty, '' AS Notes"
    AddMatlsSelectSQL += " FROM WICS_tmpmateriallistupdate"
    AddMatlsSelectSQL += " WHERE (MaterialLink_id IS NULL) AND (recStatus = 'ADD') "

    AddMatlsDoitSQL = "INSERT INTO WICS_materiallist"
    # AddMatlsDoitSQL += " (org_id, Material, Description, Plant, PartType_id,"
    AddMatlsDoitSQL += " (org_id, Material, Description, Plant,"
    AddMatlsDoitSQL += " SAPMaterialType, SAPMaterialGroup, Price, PriceUnit, Currency,"
    AddMatlsDoitSQL += " TypicalContainerQty, TypicalPalletQty, Notes)"
    AddMatlsDoitSQL += " " + AddMatlsSelectSQL
    
    app_db.session.execute(text(AddMatlsDoitSQL))
    app_db.session.commit()

    async_comm.set_async_comm_state(
        reqid,
        statecode = 'add-matl-get-recid',
        statetext = f'Getting Record ids of SAP MM60 Materials new to WICS',
        )
    UpdMaterialLinkSQL = 'UPDATE WICS_tmpmateriallistupdate, (select id, org_id, Material from WICS_materiallist) as MasterMaterials'
    UpdMaterialLinkSQL += ' set WICS_tmpmateriallistupdate.MaterialLink_id = MasterMaterials.id '
    UpdMaterialLinkSQL += ' where WICS_tmpmateriallistupdate.org_id = MasterMaterials.org_id '
    UpdMaterialLinkSQL += '   and WICS_tmpmateriallistupdate.Material = MasterMaterials.Material '
    UpdMaterialLinkSQL += "   and (MaterialLink_id IS NULL) AND (recStatus = 'ADD')"
    app_db.session.execute(text(UpdMaterialLinkSQL))
    app_db.session.commit()

    # report done and move to next step
    mandatorytaskdonekey = f'MatlX{reqid}'
    statecodeVal = ".03A."
    existingstatecode = ''
    if async_comm.async_comm_exists(mandatorytaskdonekey):
        existingstatecode = getattr(async_comm.get_async_comm_state(mandatorytaskdonekey), 'statecode', '')    # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to '' if we can't get the statecode
    MatlXval = existingstatecode + statecodeVal
    async_comm.set_async_comm_state(
        mandatorytaskdonekey,
        statecode = MatlXval,
        statetext = '',
        )
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'done-add-matl',
        statetext = f'Finished Adding SAP MM60 Materials new to WICS',
        )
    return reqid
# proc_MatlListSAPSprsheet_04_Add

def proc_MatlListSAPSprsheet_99_FinalProc(reqid):
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'done',
        statetext = 'Finished Processing Spreadsheet',
        )
    
def proc_MatlListSAPSprsheet_99_Cleanup(reqid):

    # also kill reqid, acomm, qcluster process
    keylist = [
        reqid, 
        f"{reqid}-UpdExstFldList",
        f"{reqid}-RmvMissingMatl",
        f"MatlX{reqid}",
        ]
    for key in keylist:
        async_comm.delete_async_comm(key)

    # when we can start django-q programmatically, this is where we kill that process
    # Huey is being run as always-on, so no need to kill it
    # eventually, delete this code
    # try:
    #     os.kill(int(reqid), signal.SIGTERM)
    # except AttributeError:
    #     pass
    # try:
    #     os.kill(int(reqid), signal.SIGKILL)
    # except AttributeError:
    #     pass

    # delete the temporary table
    app_db.session.query(tmpMaterialListUpdate).delete(synchronize_session=False)
    app_db.session.commit()

class PhaseEnum(Enum):
    INIT_UPL = "init-upl"
    READ_SPREADSHEET = "01ReadSpreadsheet"
    IDENTIFY_EXIST = "02IdentifyExist"
    UPDATE_EXISTING = "03UpdateExisting"
    REMOVE = "04Remove"
    ADD = "04Add"
    WANT_RESULTS = "wantresults"
    CLEANUP_FAILURE = "cleanup-after-failure"
    RESULTS_PRESENTED = "resultspresented"
    FINAL = "**FINAL**"
# PhaseEnum

@login_required
def fnUpdateMatlListfromSAP_init():
    return redirect(url_for('WICS.UpdateMatlListfromSAP'))

@login_required
def fnUpdateMatlListfromSAP():

    client_phase = request.form.get('currentPhase', None)
    client_phaseEnum = PhaseEnum(client_phase) if client_phase is not None else None
    reqid = request.form.get('reqid', None)  

    if request.method == 'POST':
        if   client_phaseEnum == PhaseEnum.INIT_UPL:
            # start Huey consumer (or at least make sure it's running) and save the pid in a cookie so we can kill it later in the cleanup proc.  When we can start Huey programmatically, this is where we start it and get the pid.
            # reqid = subprocess.Popen(
                # ['python', f'huey_consumer.py', 'app.huey -w 4']
            # ).pid
            # retinfo.set_cookie('reqid',str(reqid))

            reqid = uuid.uuid4()
            while async_comm.async_comm_exists(reqid):
                reqid = uuid.uuid4()

            UpdateExistFldList = request.form.getlist('UpIfCh')
            rmvMissingMaterial = (request.form.get('rmvMissingMaterial', False) == 'remove-missing-material')
            proc_MatlListSAPSprsheet_00InitUMLasync_comm(reqid, UpdateExistFldList, rmvMissingMaterial)
            
            retinfo = make_response(jsonify(reqid=str(reqid)))
            return retinfo

        elif client_phaseEnum == PhaseEnum.READ_SPREADSHEET:
            use_local_copy = request.form.get('use-local-copy', False) == 'use-local-copy'
            if use_local_copy:
                UMLSSName = proc_MatlListSAPSprsheet_00ResolveLocalSpreadsheetPath(reqid)
            else:
                UMLSSName = proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet(reqid)
            #endif use local copy
            proc_MatlListSAPSprsheet_01ReadSpreadsheet(reqid, UMLSSName, cleanup_file=not use_local_copy)

            acomm = async_comm.get_async_comm_state(reqid)    # something's very wrong if this doesn't exist
            acomm_dict = None if acomm is None else {c.key: getattr(acomm, c.key) for c in inspect(acomm).mapper.column_attrs}
            retinfo = make_response(jsonify(acomm_dict))
            return retinfo
        elif client_phaseEnum == PhaseEnum.IDENTIFY_EXIST:
            proc_MatlListSAPSprsheet_02_identifyexistingMaterial(reqid)
            
            acomm = async_comm.get_async_comm_state(reqid)    # something's very wrong if this doesn't exist
            acomm_dict = None if acomm is None else {c.key: getattr(acomm, c.key) for c in inspect(acomm).mapper.column_attrs}
            retinfo = make_response(jsonify(acomm_dict))
            return retinfo
        elif client_phaseEnum == PhaseEnum.UPDATE_EXISTING:
            proc_MatlListSAPSprsheet_03_UpdateExistingRecs(reqid)
            
            acomm = async_comm.get_async_comm_state(reqid)    # something's very wrong if this doesn't exist
            acomm_dict = None if acomm is None else {c.key: getattr(acomm, c.key) for c in inspect(acomm).mapper.column_attrs}
            retinfo = make_response(jsonify(acomm_dict))
            return retinfo
        elif client_phaseEnum == PhaseEnum.REMOVE:
            # skip removals for now; just go straight to the adds
            # proc_MatlListSAPSprsheet_04_Remove(reqid)
            
            acomm = async_comm.get_async_comm_state(reqid)    # something's very wrong if this doesn't exist
            acomm_dict = None if acomm is None else {c.key: getattr(acomm, c.key) for c in inspect(acomm).mapper.column_attrs}
            retinfo = make_response(jsonify(acomm_dict))
            return retinfo
        elif client_phaseEnum == PhaseEnum.ADD:
            proc_MatlListSAPSprsheet_04_Add(reqid)
            
            acomm = async_comm.get_async_comm_state(reqid)    # something's very wrong if this doesn't exist
            acomm_dict = None if acomm is None else {c.key: getattr(acomm, c.key) for c in inspect(acomm).mapper.column_attrs}
            retinfo = make_response(jsonify(acomm_dict))
            return retinfo
        elif client_phaseEnum == PhaseEnum.WANT_RESULTS:
            proc_MatlListSAPSprsheet_99_FinalProc(reqid)
            # async_comm.delete_async_comm(mandatory_commit_key)
            
            stmt = select(tmpMaterialListUpdate).where(tmpMaterialListUpdate.recStatus.startswith('err'))
            ImpErrList = app_db.session.execute(stmt).scalars().all()
            stmt = select(tmpMaterialListUpdate).where(tmpMaterialListUpdate.recStatus=='ADD')
            AddedMatlsList = app_db.session.execute(stmt).scalars().all()
            stmt = select(tmpMaterialListUpdate).where(tmpMaterialListUpdate.recStatus.startswith('DEL'))
            RemvdMatlsList = app_db.session.execute(stmt).scalars().all()
            cntext = {
                'dummyForm': FlaskForm(),       # for getting csrf_token
                'reqid': reqid,
                'ImpErrList':ImpErrList,
                'AddedMatls':AddedMatlsList,
                'RemvdMatls':RemvdMatlsList,
                }
            templt = 'Material/frmUpdateMatlListfromSAP_done.html'
            return checkTemplate_and_render(templt, **cntext)
        elif client_phaseEnum == PhaseEnum.CLEANUP_FAILURE:
            proc_MatlListSAPSprsheet_99_Cleanup(reqid)
            return make_response(jsonify(status='cleanup-complete', reqid=str(reqid)))
        elif client_phaseEnum == PhaseEnum.RESULTS_PRESENTED:
            proc_MatlListSAPSprsheet_99_Cleanup(reqid)
            return make_response(jsonify(status='results-presented', reqid=str(reqid)))
        elif client_phaseEnum == PhaseEnum.FINAL:
            return make_response(jsonify(status='final', reqid=str(reqid)))
        else:
            return make_response(jsonify(error=f'Unknown client_phase: {client_phase}'), 400)
        # endif client_phase

    else:   # req.method != 'POST'
        # (hopefully,) this is the initial phase; all others will be part of a POST request

        cntext = {
            'reqid': -1,
            }
        templt = 'Material/frmUpdateMatlListfromSAP_phase0.html'
        return checkTemplate_and_render(templt, **cntext)
    #endif req.method = 'POST'
# fnunUpdateMatlListfromSAP

# def init_UpldMatlList():
#     reqid = str(uuid.uuid4())
#     while async_comm.async_comm_exists(reqid):
#         reqid = str(uuid.uuid4())

#     UpdateExistFldList = request.form.getlist('UpIfCh')
#     rmvMissingMaterial = (request.form.get('rmvMissingMaterial', False) == 'remove-missing-material')
#     proc_MatlListSAPSprsheet_00InitUMLasync_comm(reqid, UpdateExistFldList, rmvMissingMaterial)

#     uselocalCopy = (request.form.get('use-local-copy', False) == 'use-local-copy')
#     UMLSSName = proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet(reqid, uselocalCopy)

#     pipeline = (
#         proc_MatlListSAPSprsheet_01ReadSpreadsheet.s(reqid, UMLSSName)
#         .then(proc_MatlListSAPSprsheet_02_identifyexistingMaterial.s())
#         .then(proc_MatlListSAPSprsheet_03_UpdateExistingRecs.s())
#         .then(proc_MatlListSAPSprsheet_04_Remove.s())
#         .then(proc_MatlListSAPSprsheet_04_Add.s())
#         )
#     huey.enqueue(pipeline)
    

#     acomm = async_comm.get_async_comm_state(reqid)    # something's very wrong if this doesn't exist
#     # retinfo = make_response(jsonify(reqid))
#     # return retinfo
#     return {"job_id": reqid}
# # init_UpldMatlList

# from database import HueySession
# @app.get("/SSE/UpdMatlLst/<reqid>")

