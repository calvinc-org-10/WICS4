import datetime, os, uuid
from typing import Any
from enum import Enum

from flask import (
    request, current_app,
    redirect, url_for,
    make_response, jsonify,
    )
from flask_login import login_required
from flask_wtf import FlaskForm

from sqlalchemy import (
    inspect,
    select, delete,
    )

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel, WINDOWS_EPOCH

from calvincTools.mathexpr_parser import evaluate
from calvincTools.utils import (
    coerce_date,
    ExcelWorkbook_fileext,
    checkTemplate_and_render,
    )

from database import app_db
from models import (
    UploadSAPResults, ActualCounts, MaterialList,
    async_comm,
    )

#### move to calvincTools.utils
def coerce_bool(val):
    if isinstance(val, bool):
        return val
    elif isinstance(val, str):
        val = val.strip().lower()
        if val in ['true','t','yes','y','1']:
            return True
        elif val in ['false','f','no','n','0']:
            return False
        else:
            return True
            # raise ValueError(f"Cannot coerce {val} to boolean")
    elif isinstance(val, int):
        return val != 0
        # if val == 1:
        #     return True
        # elif val == 0:
        #     return False
        # else:
        #     raise ValueError(f"Cannot coerce {val} to boolean")
    elif val is None:
        return False
    else:
        return True
        # raise ValueError(f"Cannot coerce {val} to boolean")

##############################################################
##############################################################
##############################################################

##### the suite of procs to support fnUploadCountSpreadsheet

class FatalUploadError(Exception):
    pass

def cleanupfld(fld, val, CountSprshtDateEpoch = WINDOWS_EPOCH):
    """
    fld is the name of the field in the ActualCount or MaterialList table
    val is the value to be cleaned for insertion into the fld
    Returns  {'usefld':usefld, 'cleanval': cleanval}
        usefld is a boolean indicating that val could/not be cleaned to the correct type
        cleanval is val in the correct type (if usefld==True)
    """
    cleanval = None

    if   fld == 'CountDate':
        if isinstance(val,(datetime.date, datetime.datetime)):
            usefld = True
            cleanval = val
        elif isinstance(val,int):
            usefld = True
            cleanval = from_excel(val,CountSprshtDateEpoch)
        else:
            usefld = True
            cleanval = coerce_date(val)
    elif fld in \
        ['CTD_QTY_Expr',
            ]:
        if isinstance(val,str):
            if val[0] == '=':
                val = val[1:]
        try:
            v = evaluate(str(val))
        except (SyntaxError, NameError, TypeError, ZeroDivisionError):
            v = "-- INVALID --"
        usefld = (v!="-- INVALID --")
        cleanval = str(val) if (v != "--INVALID--") else None
    elif fld in \
        ['org_id',
            ]:
        try:
            cleanval = int(val)
            usefld = True
        except:
            usefld = False
    elif fld in \
        [   'LocationOnly',
            'FLAG_PossiblyNotRecieved',
            'FLAG_MovementDuringCount',
            ]:
        val = val or 0
        try:
            cleanval = int(val)
            usefld = True
        except:
            usefld = False
    elif fld in \
        ['Material',
            'Counter',
            'LOCATION',
            'Notes',
            'TypicalContainerQty',
            'TypicalPalletQty',
            'PKGID_Desc',
            'TAGQTY',
            ]:
        usefld = (val is not None)
        if usefld: cleanval = str(val)
    else:
        usefld = True
        cleanval = val

    return {'usefld':usefld, 'cleanval': cleanval}
#end def cleanupfld

def proc_UpActCountSprsheet_00InitUpld(reqid) -> None:
    acomm = async_comm.set_async_comm_state(
        reqid,
        processname = 'Upload Counts',
        statecode = 'reading-spreadsht-init',
        statetext = 'Initializing',
        )

    # Clear lingering results, if they exist, from previous uploads
    app_db.session.execute(delete(UploadSAPResults))
    app_db.session.commit()

def proc_UpActCountSprsheet_00CopySpreadsheet(reqid) -> str:
    acomm = async_comm.set_async_comm_state(
        reqid,
        statecode = 'uploading-sprsht',
        statetext = 'Uploading Spreadsheet',
        )

    # save the file so we can open it as an excel file
    CountSprshtFile = request.files.get('CEFile')
    if CountSprshtFile is None:
        acomm = async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = 'No spreadsheet file uploaded',
            result = 'FAIL - no file',
            )
        return ""
    svdir = current_app.config.get('SAP_FILELOC', os.getcwd())
    os.makedirs(svdir, exist_ok=True)    
    fName = os.path.join(svdir, f"tmpCE{uuid.uuid4()}{ExcelWorkbook_fileext}")
    CountSprshtFile.save(fName)

    return fName

def proc_UpActCountSprsheet_01ReadSheet(reqid: Any, fName: str) -> None:
    acomm = async_comm.set_async_comm_state(
        reqid,
        statecode = 'rdng-sprsht',
        statetext = 'Reading Spreadsheet',
        )

    NOTdbFld_flags = ['**NOTdbFld**',]

    wb = load_workbook(filename=fName, read_only=True, data_only=True)
    CountSprshtDateEpoch = wb.epoch
    if 'Counts' in wb:
        ws = wb['Counts']
    else:
        acomm = async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = 'This workbook does not contain a sheet named Counts in the correct format',
            result = 'FAIL - no Counts sheet',
            )
        wb.close()
        os.remove(fName)
        return
    #endif 'Counts' in wb

    SprshtcolmnNames = ws[1]
    SprshtREQUIREDFLDS = ['Material','CountDate','Counter','LOCATION']
        # LocationOnly/CTD_QTY_Expr handled separately since at least one must be present and both can be
    SprshtcolmnMap = {}
    Sprsht_SSName_TableName_map = {
            'CountDate': 'CountDate',
            'Counter': 'Counter',
            'LOCATION': 'LOCATION',
            'org_id': 'org_id',
            'Material': 'Material',
            'LocationOnly': 'LocationOnly',
            'CTD_QTY_Expr': 'CTD_QTY_Expr',
            'Typ Cntner Qty': 'TypicalContainerQty',
            'Typ Plt Qty': 'TypicalPalletQty',
            'Notes': 'Notes',
            'PKGID_Desc': 'PKGID_Desc',
            'TAGQTY': 'TAGQTY',
            'Poss Not Rcvd': 'FLAG_PossiblyNotRecieved',
            'Mvmt Dur Ct': 'FLAG_MovementDuringCount',
            'WICSignore': NOTdbFld_flags[0],
            }
    for col in SprshtcolmnNames:
        if col.value in Sprsht_SSName_TableName_map:
            colkey = Sprsht_SSName_TableName_map[str(col.value)]
            # has this col.value already been mapped?
            if (colkey in SprshtcolmnMap and SprshtcolmnMap[colkey] is not None):
                # yes, that's a problem
                async_comm.set_async_comm_state(
                    reqid,
                    statecode = 'fatalerr',
                    statetext = f'SAP Spreadsheet has bad header row - More than one column named {col.value}.  See Calvin to fix this.',
                    result = 'FAIL - bad spreadsheet',
                    )
                wb.close()
                os.remove(fName)
                return
            else:
                assert col.column is not None, f"Column {col.value} has no column index"
                SprshtcolmnMap[colkey] = col.column - 1
            # endif previously mapped
        #endif col.value in SAP_SSName_TableName_map
    #endfor col in SAPcolmnNames

    HeaderGood = all([(reqFld in SprshtcolmnMap) for reqFld in SprshtREQUIREDFLDS])
    if not HeaderGood:
        MissingRequiredFields = [reqFld for reqFld in SprshtREQUIREDFLDS if reqFld not in SprshtcolmnMap]
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = f'Counts worksheet has bad header row - missing columns {MissingRequiredFields}.  See Calvin to fix this.',
            result = 'FAIL - bad spreadsheet',
            )
        wb.close()
        os.remove(fName)
        return

    SprshtRowNum=1
    nRowsAdded = 0
    nRowsNoMaterial = 0
    nRowsErrors = 0
    numrows = ws.max_row
    reportEveryNRows = min(100, max(1, numrows//10))

    for row in ws.iter_rows(min_row=SprshtRowNum+1, values_only=True):
        SprshtRowNum += 1
        if SprshtRowNum % reportEveryNRows == 0:
            async_comm.set_async_comm_state(
                reqid,
                statecode = 'rdng-sprsht',
                statetext = f'Reading Spreadsheet ... record {SprshtRowNum} of {numrows}<br><progress max="{ws.max_row}" value="{SprshtRowNum}"></progress>',
                )

        ignoreline = any([ 
                        all([   # you mean WICSignore -- say it !!??!!
                             NOTdbFld_flags[0] in SprshtcolmnMap, 
                             row[SprshtcolmnMap[NOTdbFld_flags[0]]]
                             ]),
                        row[SprshtcolmnMap['Material']] is None
                    ])
        if not ignoreline:
            matlnum = cleanupfld('Material', row[SprshtcolmnMap['Material']])['cleanval']
            # if no org given, check that Material unique.
            if Sprsht_SSName_TableName_map['org_id'] not in SprshtcolmnMap:
                spshtorg = None
            else:
                spshtorg = cleanupfld('org_id', row[SprshtcolmnMap['org_id']])['cleanval']
            stmt = select(MaterialList).where(MaterialList.Material==matlnum)
            matlorglist = app_db.session.execute(stmt).scalars().all()
            MatlKount = len(matlorglist)
            MatObj = None
            err_already_handled = False
            if MatlKount == 1:
                MatObj = matlorglist[0]
                spshtorg = MatObj.org_id
            if MatlKount > 1:
                if spshtorg is None:
                    resRec = UploadSAPResults(
                        errState = 'error',
                        errmsg = f"{matlnum} in multiple org_id's {tuple(matlorglist)}, but no org_id given",
                        rowNum = SprshtRowNum
                        )
                    app_db.session.add(resRec)
                    app_db.session.commit()
                    nRowsErrors += 1
                    err_already_handled = True
                else:
                    foundorg, foundrec = (rec.org_id, rec) if (rec := next((rec for rec in matlorglist if rec.org_id == spshtorg), None)) else (None, None)
                    if foundrec is not None:
                        MatObj = foundrec
                    else:
                        resRec = UploadSAPResults(
                            errState = 'error',
                            errmsg = f"{matlnum} in in multiple org_id's {tuple(rec.org_id for rec in matlorglist)}, but org_id given ({spshtorg}) is not one of them",
                            rowNum = SprshtRowNum
                            )
                        app_db.session.add(resRec)
                        app_db.session.commit()
                        nRowsErrors += 1
                        err_already_handled = True
                    # endif foundrec is not None
                #endif spshtorg is None
            #endif MatKount > 1

            if MatObj is None:
                if not err_already_handled:
                    nRowsErrors += 1
                    resRec = UploadSAPResults(
                        errState = 'error',
                        errmsg = f'either {matlnum} does not exist in MaterialList or incorrect org_id ({str(spshtorg)}) given',
                        rowNum = SprshtRowNum
                        )
                    app_db.session.add(resRec)
                    app_db.session.commit()
            else:
                rowErrs = False
                requiredFields = {reqFld: False for reqFld in SprshtREQUIREDFLDS}
                requiredFields['Both LocationOnly and CTD_QTY'] = False

                MatChanged = False
                SRec = ActualCounts()
                for fldName, colNum in SprshtcolmnMap.items():
                    if fldName in NOTdbFld_flags: continue
                    # check/correct problematic data types
                    usefld, V = cleanupfld(fldName, row[colNum], CountSprshtDateEpoch=CountSprshtDateEpoch).values()
                    if (V is not None):
                        if usefld:
                            if   fldName == 'CountDate':
                                setattr(SRec, fldName, V)
                                requiredFields['CountDate'] = True
                            elif fldName == 'Material':
                                setattr(SRec, fldName, MatObj)
                                requiredFields['Material'] = True
                            elif fldName == 'Counter':
                                setattr(SRec, fldName, V)
                                requiredFields['Counter'] = True
                            elif fldName == 'LOCATION':
                                setattr(SRec, fldName, V)
                                requiredFields['LOCATION'] = True
                            elif fldName == 'LocationOnly':
                                setattr(SRec, fldName, coerce_bool(V))
                                requiredFields['Both LocationOnly and CTD_QTY'] = True
                            elif fldName == 'CTD_QTY_Expr':
                                setattr(SRec, fldName, V)
                                requiredFields['Both LocationOnly and CTD_QTY'] = True
                            elif fldName == 'TypicalContainerQty' \
                            or fldName == 'TypicalPalletQty':
                                if V == '' or V == None: V = 0
                                if V != 0 and V != getattr(MatObj,fldName,0):
                                    setattr(MatObj, fldName, V)
                                    MatChanged = True
                            else:
                                if hasattr(SRec, fldName): setattr(SRec, fldName, V)
                            # endif fldname
                        else:   # usefld is false
                            if fldName!='CTD_QTY_Expr':
                                # we have to suspend judgement on CTD_QTY_Expr until last, because this could be a LocationOnly count
                                rowErrs = True
                                resRec = UploadSAPResults(
                                    errState = 'error',
                                    errmsg = f'{str(V)} is invalid for {fldName}',
                                    rowNum = SprshtRowNum
                                    )
                                app_db.session.add(resRec)
                                app_db.session.commit()
                        #endif usefld
                    #endif (V is not None)
                # for each column

                # now we determine if one of LocationOnly or CTD_QTY was given
                if not requiredFields['Both LocationOnly and CTD_QTY']:
                    fldName = 'CTD_QTY_Expr'
                    V = row[SprshtcolmnMap[fldName]]
                    rowErrs = True
                    resRec = UploadSAPResults(
                        errState = 'error',
                        errmsg = f'record is not marked LocationOnly and {str(V)} is invalid for {fldName}',
                        rowNum = SprshtRowNum
                        )
                    app_db.session.add(resRec)
                    app_db.session.commit()
                #endif is not LocationOnly and CTD_QTY_Expr is invalid

                # are all required fields present?
                AllRequiredPresent = True
                for keyname, Prsnt in requiredFields.items():
                    AllRequiredPresent = AllRequiredPresent and Prsnt
                    if not Prsnt:
                        rowErrs = True
                        resRec = UploadSAPResults(
                            errState = 'error',
                            errmsg = f'{keyname} missing',
                            rowNum = SprshtRowNum
                            )
                        app_db.session.add(resRec)
                        app_db.session.commit()
                    #endif keyname not Prsnt
                #endfor keyname, Prsnt in requiredFields.items()

                if not rowErrs:
                    app_db.session.add(SRec)
                    if MatChanged: app_db.session.add(MatObj)
                    app_db.session.commit()
                    resultString = str(SRec)
                    resultString += ' / LOCATION ONLY'  if SRec.LocationOnly else f' / Qty= {SRec.CTD_QTY_Expr}'
                    resultString += ' (Typ Cont Qty/Typ Plt Qty also changed)' if MatChanged else ''
                    resRec = UploadSAPResults(
                        errState = 'success',
                        errmsg = resultString,
                        rowNum = SprshtRowNum
                        )
                    app_db.session.add(resRec)
                    app_db.session.commit()
                    nRowsAdded += 1
                else:
                    nRowsErrors += 1
                #endif not rowErrs
            # endif MatObj/not MatObj
        else:
            nRowsNoMaterial += 1
        #endif not ignoreline
    # endfor row in ws.iter_rows

    resRec_Total = UploadSAPResults(
        errState = 'nRowsTotal',
        errmsg = '',
        rowNum = SprshtRowNum
        )
    resRec_Added = UploadSAPResults(
        errState = 'nRowsAdded',
        errmsg = '',
        rowNum = nRowsAdded
        )
    resRec_Errors = UploadSAPResults(
        errState = 'nRowsErrors',
        errmsg = '',
        rowNum = nRowsErrors
        )
    resRec_Ignored = UploadSAPResults(
        errState = 'nRowsIgnored',
        errmsg = '',
        rowNum = nRowsNoMaterial
        )
    app_db.session.add_all([resRec_Total, resRec_Added, resRec_Errors, resRec_Ignored])
    app_db.session.commit()

    # close and kill temp files
    wb.close()
    os.remove(fName)
# def done_UpActCountSprsheet_01ReadSheet(t):
    # report done and move to next step
    statecode = getattr(async_comm.get_async_comm_state(reqid), 'statecode', 'fatalerr')                 # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to fatalerr if we can't get the statecode
    statetext = getattr(async_comm.get_async_comm_state(reqid), 'statetext', f'No state for {reqid}')    # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to fatalerr if we can't get the statecode
    if statecode != 'fatalerr':
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'done-rdng-sprsht',
            statetext = f'Finished Reading Spreadsheet',
            )
        return reqid
    else:
        statetext = f'Error: Something went wrong while reading the spreadsheet. Please check the spreadsheet and try again. Details: {statetext}'
        # async_comm.set_async_comm_state(
        #     reqid,
        #     statecode = 'fatalerr',
        #     statetext = statetext,
        #     result = 'FAIL - error reading spreadsheet',
        #     )
        raise FatalUploadError(statetext)
# proc_UpActCountSprsheet_01ReadSheet

def proc_UpActCountSprsheet_99_FinalProc(reqid:Any) -> None:
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'done',
        statetext = 'Finished Processing Spreadsheet',
        )

def proc_UpActCountSprsheet_99_Cleanup(reqid:Any) -> None:

    # also kill reqid, acomm, qcluster process
    keylist = [
        reqid, 
        ]
    for key in keylist:
        async_comm.delete_async_comm(key)

    # try:
    #     os.kill(int(reqid), signal.SIGTERM)
    # except AttributeError:
    #     pass
    # try:
    #     os.kill(int(reqid), signal.SIGKILL)
    # except AttributeError:
    #     pass

    # delete the temporary table
    app_db.session.execute(delete(UploadSAPResults))
# proc_UpActCountSprsheet_99_Cleanup

class PhaseEnum(Enum):
    INIT_UPL = "init-upl"
    READ_SPREADSHEET = "01ReadSpreadsheet"

    WANT_RESULTS = "wantresults"
    CLEANUP_FAILURE = "cleanup-after-failure"
    RESULTS_PRESENTED = "resultspresented"
    FINAL = "**FINAL**"
# PhaseEnum

@login_required
def fnUploadActCountSprsht_init():
    return redirect(url_for('WICS.UploadActualCounts'))

@login_required
def fnUploadActCountSprsht():

    client_phase = request.form.get('currentPhase', None)
    client_phaseEnum = PhaseEnum(client_phase) if client_phase is not None else None
    reqid = request.form.get('reqid', None)  

    if request.method == 'POST':
        if   client_phaseEnum == PhaseEnum.INIT_UPL:
            reqid = str(uuid.uuid4())
            while async_comm.async_comm_exists(reqid):
                reqid = str(uuid.uuid4())
                
            proc_UpActCountSprsheet_00InitUpld(reqid)

            retinfo = make_response(jsonify(reqid=str(reqid)))
            return retinfo
        elif client_phaseEnum == PhaseEnum.READ_SPREADSHEET:
            # save the file so we can open it as an excel file
            fName = proc_UpActCountSprsheet_00CopySpreadsheet(reqid)

            proc_UpActCountSprsheet_01ReadSheet(reqid, fName)

            acomm = async_comm.get_async_comm_state(reqid)    # something's very wrong if this doesn't exist
            acomm_dict = None if acomm is None else {c.key: getattr(acomm, c.key) for c in inspect(acomm).mapper.column_attrs}
            retinfo = make_response(jsonify(acomm_dict))
            return retinfo
        elif client_phaseEnum == PhaseEnum.WANT_RESULTS:
            stmt = select(UploadSAPResults).where(UploadSAPResults.errState.in_(['nRowsTotal','nRowsAdded','nRowsErrors','nRowsIgnored']))
            QSet = app_db.session.execute(stmt).scalars().all()
            SprshtRowNum =      getattr([rec for rec in QSet if rec.errState == 'nRowsTotal'][0],'rowNum', 0)
            nRowsAdded =        getattr([rec for rec in QSet if rec.errState == 'nRowsAdded'][0],'rowNum', 0)
            nRowsErrors =       getattr([rec for rec in QSet if rec.errState == 'nRowsErrors'][0],'rowNum', 0)
            nRowsNoMaterial =   getattr([rec for rec in QSet if rec.errState == 'nRowsIgnored'][0],'rowNum', 0)

            stmt = select(UploadSAPResults).where(UploadSAPResults.errState.not_in(['nRowsTotal','nRowsAdded','nRowsErrors','nRowsIgnored'])).order_by(UploadSAPResults.rowNum)
            UplResults = app_db.session.execute(stmt).scalars().all()
            cntext = {
                'dummyForm': FlaskForm(),       # for getting csrf_token
                'reqid': reqid,
                'UplResults':UplResults,
                'ResultStats': {
                    'nRowsRead': SprshtRowNum - 1,
                        # -1 because header doesn't count
                    'nRowsAdded': nRowsAdded ,
                    'nRowsNoMaterial': nRowsNoMaterial,
                    'nRowsErrors': nRowsErrors,
                    },
                }
            templt = 'ActualCounts/frm_uploadCountEntry_Success.html'
            return checkTemplate_and_render(templt, **cntext)
        elif client_phaseEnum == PhaseEnum.CLEANUP_FAILURE:
            proc_UpActCountSprsheet_99_Cleanup(reqid)
            return make_response(jsonify(status='cleanup-complete', reqid=str(reqid)))
        elif client_phaseEnum == PhaseEnum.RESULTS_PRESENTED:
            proc_UpActCountSprsheet_99_Cleanup(reqid)
            return make_response(jsonify(status='results-presented', reqid=str(reqid)))
        elif client_phaseEnum == PhaseEnum.FINAL:
            return make_response(jsonify(status='final', reqid=str(reqid)))
        else:
            return make_response(jsonify(error=f'Unknown client_phase: {client_phase}'), 400)
        # endif client_phaseEnum
    else:   # req.method != 'POST'
        cntext = {
                }
        templt = 'ActualCounts/frm_UploadCountEntrySprdsht.html'
    #endif req.method

    return checkTemplate_and_render(templt, **cntext)
# fnUploadActCountSprsht
