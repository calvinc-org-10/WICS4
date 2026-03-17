import uuid, os, re as regex, ast, json, time
from functools import wraps

from flask import (
    current_app,
    session, 
    request, make_response, Response, 
    jsonify, stream_with_context,
    )
from flask_login import login_required

from sqlalchemy import text
from sqlalchemy.sql import select

from openpyxl import load_workbook

from async_tasks import huey

from calvincTools.utils import (
    checkTemplate_and_render,
    ExcelWorkbook_fileext,
    )

from calvincTools.models import (cParameters, )
from app import app as thisapp
from database import (app_db,)
from models import (
    tmpMaterialListUpdate, SAPPlants_org,
    async_comm
    )
    

####################################################################################
####################################################################################
####################################################################################

##### the suite of procs to support fnUpdateMatlListfromSAP

class FatalUploadError(Exception):
    pass


def huey_task_with_app_context(fn):
    """Ensure each Huey task execution gets a fresh Flask app context."""

    @wraps(fn)
    def wrapped(*args, **kwargs):
        with thisapp.app_context():
            return fn(*args, **kwargs)

    return wrapped

def proc_MatlListSAPSprsheet_00InitUMLasync_comm(reqid, UpdateExistFldList, rmvMissingMaterial=False):
    # these first calls should create the async_comm record with pk=reqid.  All subsequent calls will update that same record until we delete it in the cleanup proc at the end.
    acomm = async_comm.set_async_comm_state(
        reqid,
        statecode = 'rdng-sprsht-init',
        statetext = 'Initializing ...',
        )   
    async_comm.set_async_comm_state(
        f'{reqid}-UpdExstFldList',
        statecode = 'UpdateExistFldList',
        statetext = f'{UpdateExistFldList}',
        )
    async_comm.set_async_comm_state(
        f'{reqid}-RmvMissingMatl',
        statecode = 'RmvMissingMatl',
        statetext = f'{rmvMissingMaterial}',
        )

def proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet(reqid, uselocalCopy=False):
    acomm = async_comm.set_async_comm_state(
        reqid,
        statecode = 'uploading-sprsht',
        statetext = 'Uploading Spreadsheet',
        )

    SAPFile = request.files.get('SAPFile')
    if SAPFile is None:
        acomm = async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = 'No file uploaded. Please upload an Excel spreadsheet and try again.',
            result = 'FAIL - no file uploaded',
            )
        return
    svdir = cParameters.get_parameter('SAP-FILELOC') if not uselocalCopy else ''
    fName = svdir+"tmpMatlList"+str(reqid)+ExcelWorkbook_fileext
    SAPFile.save(fName)

    return fName

@huey.task()
@huey_task_with_app_context
def proc_MatlListSAPSprsheet_01ReadSpreadsheet(reqid, fName):
    acomm = async_comm.set_async_comm_state(
        reqid,
        statecode = 'rdng-sprsht',
        statetext = 'Reading Spreadsheet',
        )

    if len(fName)<1 or not os.path.exists(fName):
        statetext = f'Spreadsheet file {fName} not found. Please try again.'
        acomm = async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = statetext,
            result = 'FAIL - file not found',
            )
        raise FatalUploadError(statetext)
    
    app_db.session.query(tmpMaterialListUpdate).delete(synchronize_session=False)
    app_db.session.commit()
    
    wb = load_workbook(filename=fName, read_only=True)
    ws = wb.active
    assert ws is not None, "Error: Spreadsheet appears to be blank. Please fix this and try again."
    SAPcolmnNames = ws[1]
    SAPcol:dict[str, int|None] = {'Plant':None,'Material': None}
    SAP_SSName_TableName_map = {
            'Material': 'Material',
            'Material description': 'Description',
            'Plant': 'Plant', 'Plnt': 'Plant',
            'Material type': 'SAPMaterialType',  'MTyp': 'SAPMaterialType',
            'Material Group': 'SAPMaterialGroup', 'Matl Group': 'SAPMaterialGroup',
            'Manufact.': 'SAPManuf', 
            'MPN': 'SAPMPN', 
            'ABC': 'SAPABC', 
            'Price': 'Price', 'Standard price': 'Price',
            'Price unit': 'PriceUnit', 'per': 'PriceUnit',
            'Currency':'Currency',
            }
    for col in SAPcolmnNames:
        assert col.value is not None, f"Error: Blank column header found in spreadsheet at column {col.column}. Please fix this and try again."
        assert col.column is not None, f"Error: Column with blank header has no column number. This shouldn't happen. Please check the spreadsheet and try again."
        if col.value in SAP_SSName_TableName_map:
            colval = str(col.value)
            SAPcol[SAP_SSName_TableName_map[colval]] = col.column - 1
    if (SAPcol['Material'] == None or SAPcol['Plant'] == None):
        statetext = 'SAP Spreadsheet has bad header row. Plant and/or Material is missing.  See Calvin to fix this.'
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'fatalerr',
            statetext = statetext,
            result = 'FAIL - bad spreadsheet',
            )

        wb.close()
        os.remove(fName)
        raise FatalUploadError(statetext)

    numrows = ws.max_row
    nRows = 0
    reportEveryNRows = min(100, max(1, numrows//10))
    for row in ws.iter_rows(min_row=2, values_only=True):
        nRows += 1
        if nRows % reportEveryNRows == 0:
            async_comm.set_async_comm_state(
                reqid,
                statecode = 'rdng-sprsht',
                statetext = f'Reading Spreadsheet ... record {nRows} of {numrows}<br><progress max="{numrows}" value="{nRows}"></progress>',
                )

        if row[SAPcol['Material']]==None: MatNum = ''
        else: MatNum = row[SAPcol['Material']]
        validTmpRec = False
        ## create a blank tmpMaterialListUpdate record,
        newrec = tmpMaterialListUpdate()
        if regex.match(".*[\n\t\xA0].*",str(MatNum)):
            validTmpRec = True
            ## refuse to work with special chars embedded in the MatNum
            newrec.recStatus = 'err-MatlNum'
            newrec.errmsg = f'error: {MatNum!a} is an unusable part number. It contains invalid characters and cannot be added to WICS'
        elif len(str(MatNum)):
            validTmpRec = True
            plant_col = SAPcol['Plant']
            if plant_col is not None:
                org_row = SAPPlants_org.query.filter_by(SAPPlant=row[plant_col]).first()
                if org_row is not None:
                    newrec.org = org_row.org
        # endif invalid Material
        if validTmpRec:
            ## populate by looping through SAPcol,
            ## then save
            for dbColName, ssColNum in SAPcol.items():
                assert ssColNum is not None, f"Error: Column {dbColName} has no column number. This shouldn't happen. Please check the spreadsheet and try again."
                setattr(newrec,dbColName,row[ssColNum])
            
            app_db.session.add(newrec)
            app_db.session.commit()
    # endfor

    wb.close()
    os.remove(fName)

    # report done and move to next step
    statecode = getattr(async_comm.get_async_comm_state(reqid), 'statecode', 'fatalerr')                 # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to fatalerr if we can't get the statecode
    statetext = getattr(async_comm.get_async_comm_state(reqid), 'statetext', f'No state for {reqid}')    # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to fatalerr if we can't get the statecode
    if statecode != 'fatalerr':
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'done-rdng-sprsht',
            statetext = f'Finished Reading Spreadsheet',
            )
        return reqid
    else:
        statetext = f'Error: Something went wrong while reading the spreadsheet. Please check the spreadsheet and try again. Details: {statetext}'
        # async_comm.set_async_comm_state(
        #     reqid,
        #     statecode = 'fatalerr',
        #     statetext = statetext,
        #     result = 'FAIL - error reading spreadsheet',
        #     )
        raise FatalUploadError(statetext)
# proc_MatlListSAPSprsheet_01ReadSpreadsheet

@huey.task()
@huey_task_with_app_context
def proc_MatlListSAPSprsheet_02_identifyexistingMaterial(reqid):
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'get-matl-link',
        statetext = f'Finding SAP MM60 Materials already in WICS Material List',
        )
    UpdMaterialLinkSQL = 'UPDATE WICS_tmpmateriallistupdate, (select id, org_id, Material from WICS_materiallist) as MasterMaterials'
    UpdMaterialLinkSQL += ' set WICS_tmpmateriallistupdate.MaterialLink_id = MasterMaterials.id, '
    UpdMaterialLinkSQL += "     WICS_tmpmateriallistupdate.recStatus = 'FOUND' "
    UpdMaterialLinkSQL += ' where WICS_tmpmateriallistupdate.org_id = MasterMaterials.org_id '
    UpdMaterialLinkSQL += '   and WICS_tmpmateriallistupdate.Material = MasterMaterials.Material '
    app_db.session.execute(text(UpdMaterialLinkSQL))
    app_db.session.commit()

    async_comm.set_async_comm_state(
        reqid,
        statecode = 'id-del-matl',
        statetext = f'Identifying WICS Materials no longer in SAP MM60 Materials',
        )
    MustKeepMatlsSelCond = ''
    MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
    MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT tmucopy.MaterialLink_id AS Material_id FROM WICS_tmpmateriallistupdate tmucopy WHERE tmucopy.MaterialLink_id IS NOT NULL)'
    MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
    MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_actualcounts)'
    MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
    MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_countschedule)'
    MustKeepMatlsSelCond += ' AND ' if MustKeepMatlsSelCond else ''
    MustKeepMatlsSelCond += 'id NOT IN (SELECT DISTINCT Material_id FROM WICS_sap_sohrecs)'

    DeleteMatlsSelectSQL = "INSERT INTO WICS_tmpmateriallistupdate (recStatus, delMaterialLink, MaterialLink_id, org_id, Material, Description, Plant "
    DeleteMatlsSelectSQL += ", SAPMaterialType, SAPMaterialGroup, Currency  ) "    # these can go once I set null=True on these fields
    DeleteMatlsSelectSQL += " SELECT  concat('DEL ',FORMAT(id,0)), id, NULL, org_id, Material, Description, Plant "
    DeleteMatlsSelectSQL += ", SAPMaterialType, SAPMaterialGroup, Currency  "    # these can go once I set null=True on these fields
    DeleteMatlsSelectSQL += " FROM WICS_materiallist"
    DeleteMatlsSelectSQL += f" WHERE ({MustKeepMatlsSelCond})"
    app_db.session.execute(text(DeleteMatlsSelectSQL))
    app_db.session.commit()

    async_comm.set_async_comm_state(
        reqid,
        statecode = 'id-add-matl',
        statetext = f'Identifying SAP MM60 Materials new to WICS',
        )
    MarkAddMatlsSelectSQL = "UPDATE WICS_tmpmateriallistupdate"
    MarkAddMatlsSelectSQL += " SET recStatus = 'ADD'"
    MarkAddMatlsSelectSQL += " WHERE (MaterialLink_id IS NULL) AND (recStatus is NULL)"
    app_db.session.execute(text(MarkAddMatlsSelectSQL))
    app_db.session.commit()

    # report done and move to next step
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'get-matl-link-done',
        statetext = f'Finished linking SAP MM60 list to existing WICS Materials',
        )
    
    return reqid
# proc_MatlListSAPSprsheet_02_identifyexistingMaterial

@huey.task()
@huey_task_with_app_context
def proc_MatlListSAPSprsheet_03_UpdateExistingRecs(reqid):
    def setstate_MatlListSAPSprsheet_03_UpdateExistingRecs(fldName):
        acomm = async_comm.set_async_comm_state(
            reqid,
            statecode = 'upd-existing-recs',
            statetext = f'Updating _{fldName}_ Field in Existing Records',
            )

    setstate_MatlListSAPSprsheet_03_UpdateExistingRecs('')

    # (Form Name, db fld Name, zero/blank value)
    FormTodbFld_map = [
        ('Description','Description','""'),
        ('SAPMatlType','SAPMaterialType','""'),
        ('SAPMatlGroup','SAPMaterialGroup','""'),
        ('SAPManuf','SAPManuf','""'),
        ('SAPMPN','SAPMPN','""'),
        ('SAPABC','SAPABC','""'),
        ('SAPPrice','Price',0),
        ('SAPPrice','PriceUnit',0),
        ('SAPPrice','Currency','""'),
    ]

    UpdateExistFldList_str = getattr(async_comm.get_async_comm_state(f"{reqid}-UpdExstFldList"), 'statetext', '[]')    # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to '' if we can't get the statetext
    UpdateExistFldList = ast.literal_eval(UpdateExistFldList_str)

    if UpdateExistFldList:
        for formName, dbName, zeroVal in FormTodbFld_map:
            if formName in UpdateExistFldList:
                setstate_MatlListSAPSprsheet_03_UpdateExistingRecs(dbName)
                # UPDATE this field
                UpdSQLSetStmt = f"MatlList.{dbName}=tmpMatl.{dbName}"
                UpdSQLWhereStmt = f"(IFNULL(tmpMatl.{dbName},{zeroVal}) != {zeroVal} AND IFNULL(MatlList.{dbName},{zeroVal})!=IFNULL(tmpMatl.{dbName},{zeroVal}))"

                UpdSQLStmt = "UPDATE WICS_materiallist AS MatlList, WICS_tmpmateriallistupdate AS tmpMatl"
                UpdSQLStmt += f" SET {UpdSQLSetStmt}"
                UpdSQLStmt += f" WHERE (tmpMatl.MaterialLink_id=MatlList.id) AND {UpdSQLWhereStmt}"
                app_db.session.execute(text(UpdSQLStmt))
                app_db.session.commit()
            #endif formName in UpdateExistFldList
        #endfor
    # endif UpdateExistFldList not empty

    # report done and move to next step
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'upd-existing-recs-done',
        statetext = f'Finished Updating Existing Records to MM60 values',
        )
    return reqid
# proc_MatlListSAPSprsheet_03_UpdateExistingRecs

@huey.task()
@huey_task_with_app_context
def proc_MatlListSAPSprsheet_04_Remove(reqid):
    ## MustKeepMatlsDelCond = ''
    ## if MustKeepMatlsDelCond: MustKeepMatlsDelCond += ' AND '
    ## MustKeepMatlsDelCond += 'id IN (SELECT DISTINCT delMaterialLink FROM WICS_tmpmateriallistupdate WHERE recStatus like "DEL%")'

    doRmv_str = getattr(async_comm.get_async_comm_state(f"{reqid}-RmvMissingMatl"), 'statetext', 'False')
    doRmv = ast.literal_eval(doRmv_str)

    if not doRmv:
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'del-matl-skip',
            statetext = f'Not Removing WICS Materials no longer in SAP MM60 Materials',
            )
        proc_MatlListSAPSprsheet_04_Add(reqid)
        return
    else:
        async_comm.set_async_comm_state(
            reqid,
            statecode = 'del-matl',
            statetext = f'Removing WICS Materials no longer in SAP MM60 Materials',
            )

        # do the Removals
        ## DeleteMatlsDoitSQL = "DELETE FROM WICS_materiallist"
        ## DeleteMatlsDoitSQL += f" WHERE ({MustKeepMatlsDelCond})"
        DeleteMatlsDoitSQL = 'DELETE MATL'
        DeleteMatlsDoitSQL += ' FROM WICS_materiallist AS MATL INNER JOIN WICS_tmpmateriallistupdate AS TMP'
        DeleteMatlsDoitSQL += '    ON MATL.id = TMP.delMaterialLink'
        DeleteMatlsDoitSQL += ' WHERE TMP.recStatus like "DEL%"'
        app_db.session.execute(text(DeleteMatlsDoitSQL))
        app_db.session.commit()
    # endif doRmv

    # report done and move to next step    
    mandatorytaskdonekey = f'MatlX{reqid}'
    statecodeVal = ".03D."
    existingstatecode = ''
    if async_comm.async_comm_exists(mandatorytaskdonekey): 
        existingstatecode = getattr(async_comm.get_async_comm_state(mandatorytaskdonekey), 'statecode', '')    # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to '' if we can't get the statecode
    MatlXval = existingstatecode + statecodeVal
    async_comm.set_async_comm_state(
        mandatorytaskdonekey,
        statecode = MatlXval,
        statetext = '',
        )
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'done-del-matl',
        statetext = 'Finished Removing' if doRmv else 'Skipped Removal of' + 'WICS Materials no longer in SAP MM60 Materials',
        )
    return reqid
# proc_MatlListSAPSprsheet_04_Remove
@huey.task()
@huey_task_with_app_context
def proc_MatlListSAPSprsheet_04_Add(reqid):
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'add-matl',
        statetext = f'Adding SAP MM60 Materials new to WICS',
        )
    # phase out the UNKNOWN type; just leave WhsePartType_id blank for new records until we can get a real type assigned in WICS
    # UnknownTypeID = WhsePartTypes.objects.using(dbToUse).get(WhsePartType=WICS.globals._PartTypeName_UNKNOWN)

    # do the adds
    # one day django will implement insert ... select.  Until then ...
    # come back to this one day and rewrite it to use the ORM instead of raw SQL;
    AddMatlsSelectSQL = "SELECT"
    # AddMatlsSelectSQL += " org_id, Material, Description, Plant, " + str(UnknownTypeID.pk) + " AS PartType_id,"
    AddMatlsSelectSQL += " org_id, Material, Description, Plant,"
    AddMatlsSelectSQL += " SAPMaterialType, SAPMaterialGroup, Price, PriceUnit, Currency,"
    AddMatlsSelectSQL += " '' AS TypicalContainerQty, '' AS TypicalPalletQty, '' AS Notes"
    AddMatlsSelectSQL += " FROM WICS_tmpmateriallistupdate"
    AddMatlsSelectSQL += " WHERE (MaterialLink_id IS NULL) AND (recStatus = 'ADD') "

    AddMatlsDoitSQL = "INSERT INTO WICS_materiallist"
    # AddMatlsDoitSQL += " (org_id, Material, Description, Plant, PartType_id,"
    AddMatlsDoitSQL += " (org_id, Material, Description, Plant,"
    AddMatlsDoitSQL += " SAPMaterialType, SAPMaterialGroup, Price, PriceUnit, Currency,"
    AddMatlsDoitSQL += " TypicalContainerQty, TypicalPalletQty, Notes)"
    AddMatlsDoitSQL += " " + AddMatlsSelectSQL
    
    app_db.session.execute(text(AddMatlsDoitSQL))
    app_db.session.commit()

    async_comm.set_async_comm_state(
        reqid,
        statecode = 'add-matl-get-recid',
        statetext = f'Getting Record ids of SAP MM60 Materials new to WICS',
        )
    UpdMaterialLinkSQL = 'UPDATE WICS_tmpmateriallistupdate, (select id, org_id, Material from WICS_materiallist) as MasterMaterials'
    UpdMaterialLinkSQL += ' set WICS_tmpmateriallistupdate.MaterialLink_id = MasterMaterials.id '
    UpdMaterialLinkSQL += ' where WICS_tmpmateriallistupdate.org_id = MasterMaterials.org_id '
    UpdMaterialLinkSQL += '   and WICS_tmpmateriallistupdate.Material = MasterMaterials.Material '
    UpdMaterialLinkSQL += "   and (MaterialLink_id IS NULL) AND (recStatus = 'ADD')"
    app_db.session.execute(text(UpdMaterialLinkSQL))
    app_db.session.commit()

    # report done and move to next step
    mandatorytaskdonekey = f'MatlX{reqid}'
    statecodeVal = ".03A."
    existingstatecode = ''
    if async_comm.async_comm_exists(mandatorytaskdonekey):
        existingstatecode = getattr(async_comm.get_async_comm_state(mandatorytaskdonekey), 'statecode', '')    # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to '' if we can't get the statecode
    MatlXval = existingstatecode + statecodeVal
    async_comm.set_async_comm_state(
        mandatorytaskdonekey,
        statecode = MatlXval,
        statetext = '',
        )
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'done-add-matl',
        statetext = f'Finished Adding SAP MM60 Materials new to WICS',
        )
    return reqid
# proc_MatlListSAPSprsheet_04_Add

def proc_MatlListSAPSprsheet_99_FinalProc(reqid):
    async_comm.set_async_comm_state(
        reqid,
        statecode = 'done',
        statetext = 'Finished Processing Spreadsheet',
        )
    
def proc_MatlListSAPSprsheet_99_Cleanup(reqid):
    # also kill reqid, acomm, qcluster process
    async_comm.delete_async_comm(reqid)
    async_comm.delete_async_comm(f"{reqid}-UpdExstFldList")

    # when we can start django-q programmatically, this is where we kill that process
    # Huey is being run as always-on, so no need to kill it
    # eventually, delete this code
    # try:
    #     os.kill(int(reqid), signal.SIGTERM)
    # except AttributeError:
    #     pass
    # try:
    #     os.kill(int(reqid), signal.SIGKILL)
    # except AttributeError:
    #     pass

    # delete the temporary table
    app_db.session.query(tmpMaterialListUpdate).delete(synchronize_session=False)
    app_db.session.commit()

@login_required
def fnUpdateMatlListfromSAP():

    client_phase = request.form.get('phase', None)
    reqid = request.cookies.get('reqid', None)  # where's thee reqid kept?  # eventually, delete the use of cookies for this and just pass the reqid back and forth in the request body or something; using cookies was just a quick way to get it working without having to change the frontend code much, but it's not ideal since it relies on the client to keep track of the reqid correctly and send it back with each request, which could lead to issues if the client doesn't do that correctly.  Passing the reqid back and forth in the request body would be more reliable and easier to manage, but would require changes to the frontend code to include the reqid in each request.

    if request.method == 'POST':
        # check if the mandatory commits have been done and change the status code if so
        if reqid is not None:
            mandatory_commit_key = f'MatlX{reqid}'
            mandatory_commit_list = ['03A', '03D']
            if async_comm.async_comm_exists(mandatory_commit_key):
                mandatory_commits_recorded = async_comm.get_async_comm_state(mandatory_commit_key).statecode # type: ignore
                if all((c in str(mandatory_commits_recorded)) for c in mandatory_commit_list):
                    proc_MatlListSAPSprsheet_99_FinalProc(reqid)
                    async_comm.delete_async_comm(mandatory_commit_key)

        if client_phase=='init-upl':
            # start Huey consumer (or at least make sure it's running) and save the pid in a cookie so we can kill it later in the cleanup proc.  When we can start Huey programmatically, this is where we start it and get the pid.
            # reqid = subprocess.Popen(
                # ['python', f'huey_consumer.py', 'app.huey -w 4']
            # ).pid
            # retinfo.set_cookie('reqid',str(reqid))

            reqid = uuid.uuid4()
            while async_comm.async_comm_exists(reqid):
                reqid = uuid.uuid4()

            UpdateExistFldList = request.form.getlist('UpIfCh')
            proc_MatlListSAPSprsheet_00InitUMLasync_comm(reqid, UpdateExistFldList)

            if request.form.get('use-local-copy', False) == 'use-local-copy':
                UMLSSName = request.files.get('SAPFile').filename # type: ignore
            else:
                UMLSSName = proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet(reqid)
            #endif use local copy
            proc_MatlListSAPSprsheet_01ReadSpreadsheet(reqid, UMLSSName)

            acomm = async_comm.get_async_comm_state(reqid)    # something's very wrong if this doesn't exist
            retinfo = make_response(jsonify(acomm))
            # retinfo.set_cookie('reqid',str(reqid))
            return retinfo
        elif client_phase=='waiting':
            acomm = async_comm.get_async_comm_state(reqid)    # something's very wrong if this doesn't exist
            retinfo = make_response(jsonify(acomm))
            return retinfo
        elif client_phase=='wantresults':
            stmt = select(tmpMaterialListUpdate).where(tmpMaterialListUpdate.recStatus.startswith('err'))
            ImpErrList = app_db.session.execute(stmt).mappings().all()
            stmt = select(tmpMaterialListUpdate).where(tmpMaterialListUpdate.recStatus=='ADD')
            AddedMatlsList = app_db.session.execute(stmt).mappings().all()
            stmt = select(tmpMaterialListUpdate).where(tmpMaterialListUpdate.recStatus.startswith('DEL'))
            RemvdMatlsList = app_db.session.execute(stmt).mappings().all()
            cntext = {
                'ImpErrList':ImpErrList,
                'AddedMatls':AddedMatlsList,
                'RemvdMatls':RemvdMatlsList,
                }
            templt = 'Material/frmUpdateMatlListfromSAP_done.html'
            return checkTemplate_and_render(templt, cntext)
        elif client_phase=='cleanup-after-failure':
            pass
        elif client_phase=='resultspresented':
            proc_MatlListSAPSprsheet_99_Cleanup(reqid)
            retinfo = make_response(jsonify(success=True))
            # retinfo.delete_cookie('reqid')

            return retinfo
        else:
            return
        #endif client_phase
    else:   # req.method != 'POST'
        # (hopefully,) this is the initial phase; all others will be part of a POST request

        cntext = {
            'reqid': -1,
            }
        templt = 'Material/frmUpdateMatlListfromSAP_phase0.html'
        return checkTemplate_and_render(templt, **cntext)
    #endif req.method = 'POST'
# fnunUpdateMatlListfromSAP

def init_UpldMatlList():
    reqid = str(uuid.uuid4())
    while async_comm.async_comm_exists(reqid):
        reqid = str(uuid.uuid4())

    UpdateExistFldList = request.form.getlist('UpIfCh')
    rmvMissingMaterial = (request.form.get('rmvMissingMaterial', False) == 'remove-missing-material')
    proc_MatlListSAPSprsheet_00InitUMLasync_comm(reqid, UpdateExistFldList, rmvMissingMaterial)

    uselocalCopy = (request.form.get('use-local-copy', False) == 'use-local-copy')
    UMLSSName = proc_MatlListSAPSprsheet_00CopyUMLSpreadsheet(reqid, uselocalCopy)

    pipeline = (
        proc_MatlListSAPSprsheet_01ReadSpreadsheet.s(reqid, UMLSSName)
        .then(proc_MatlListSAPSprsheet_02_identifyexistingMaterial.s())
        .then(proc_MatlListSAPSprsheet_03_UpdateExistingRecs.s())
        .then(proc_MatlListSAPSprsheet_04_Remove.s())
        .then(proc_MatlListSAPSprsheet_04_Add.s())
        )
    huey.enqueue(pipeline)
    

    acomm = async_comm.get_async_comm_state(reqid)    # something's very wrong if this doesn't exist
    # retinfo = make_response(jsonify(reqid))
    # return retinfo
    return {"job_id": reqid}
# init_UpldMatlList

# @app.get("/SSE/ENDUpdMatlLst/<reqid>")
def closeup_UpldMatlList(reqid):
    ...
        # elif client_phase=='wantresults':
        #     stmt = select(tmpMaterialListUpdate).where(tmpMaterialListUpdate.recStatus.startswith('err'))
        #     ImpErrList = app_db.session.execute(stmt).mappings().all()
        #     stmt = select(tmpMaterialListUpdate).where(tmpMaterialListUpdate.recStatus=='ADD')
        #     AddedMatlsList = app_db.session.execute(stmt).mappings().all()
        #     stmt = select(tmpMaterialListUpdate).where(tmpMaterialListUpdate.recStatus.startswith('DEL'))
        #     RemvdMatlsList = app_db.session.execute(stmt).mappings().all()
        #     cntext = {
        #         'ImpErrList':ImpErrList,
        #         'AddedMatls':AddedMatlsList,
        #         'RemvdMatls':RemvdMatlsList,
        #         }
        #     templt = 'Material/frmUpdateMatlListfromSAP_done.html'
        #     return checkTemplate_and_render(templt, cntext)
        # elif client_phase=='cleanup-after-failure':
        #     pass
        # elif client_phase=='resultspresented':
        #     proc_MatlListSAPSprsheet_99_Cleanup(reqid)
        #     retinfo = make_response(jsonify(success=True))
        #     # retinfo.delete_cookie('reqid')

        #     return retinfo
# closeup_UpldMatlList

# from database import HueySession
# @app.get("/SSE/UpdMatlLst/<reqid>")
def progress_UpdML(reqid):

    def generate():
        last_version = 0

        while True:
            # session = HueySession()

            row = async_comm.get_async_comm_state(reqid)    # if the record has been deleted (e.g. by cleanup after failure), this will throw an exception, so default to None if we can't get the record

            if row and row.version > last_version:

                payload = json.dumps({
                    "statecode": row.statecode,
                    "statetext": row.statetext
                })  #should I dump the whole record here instead of just statecode and statetext?  Maybe not a good idea if there are big text fields or something, but it would be more flexible for the frontend if it had access to all the fields without me having to predict which ones it might want.  For now, I'll just include statecode and statetext since those are the ones I know the frontend will need, and I can always add more later if needed.

                yield f"data: {payload}\n\n"

                last_version = row.version

                if row.statecode == "done":
                    break

            # session.close()

            yield ": keepalive\n\n"

            time.sleep(1)
        # endwhile (until we break on statecode == "done")
    # generate

    r = Response(stream_with_context(generate()),
                 mimetype="text/event-stream")

    r.headers["X-Accel-Buffering"] = "no"

    return r

def PLACEHOLDER_fnUpdateMatlListfromSAP():
    # This is a placeholder for the function that will update the material list from SAP.
    # The actual implementation will depend on how you plan to connect to SAP and retrieve the data.
    # For now, it simply returns a message indicating that the function was called. You can replace this with the actual logic to update the material list from SAP.
    return "Material list update from SAP function called."
# fnUpdateMatlListfromSAP
