import os, uuid, re as regex
import ast
import subprocess, signal
import json
from functools import partial

from sqlalchemy import select

from flask import (
    request, session, 
    jsonify,
    current_app,
    )
from flask_login import login_required, current_user

from datetime import date

from openpyxl import load_workbook

from calvincTools.utils import (
    calvindate, 
    ExcelWorkbook_fileext,
    checkTemplate_and_render,
    coerce_date,
    )

from models import (
    SAP_SOHRecs, SAPPlants_org, UnitsOfMeasure, UploadSAPResults,
    #, VIEW_SAP
    WhsePartTypes, MaterialList, tmpMaterialListUpdate,
    async_comm,
    )

from database import app_db


def nearestSAPDate(for_date=date.today()) -> date|None:
    """
    returns the nearest SAP_SOHRecs.uploaded_at date that is less than or equal to for_date
    if no SAP_SOHRecs exist, returns None
    """
    
    stmt = select(SAP_SOHRecs.uploaded_at).where(SAP_SOHRecs.uploaded_at <= for_date).order_by(SAP_SOHRecs.uploaded_at.desc()).limit(1)
    nearest_date = app_db.session.execute(stmt).scalar_one_or_none()
    
    if nearest_date is None:
        # if there are any SAP_SOHRecs, return the earliest one, even if it's after for_date
        stmt_earliest = select(SAP_SOHRecs.uploaded_at).order_by(SAP_SOHRecs.uploaded_at.asc()).limit(1)
        nearest_date = app_db.session.execute(stmt_earliest).scalar_one_or_none()
    
    return nearest_date
# nearestSAPDate

def fnSAPExists(reqDate:date=date.today()) -> bool:
    """
    returns true or false indicating if SAP_SOH data exists for reqDate
    """

    return nearestSAPDate(for_date=reqDate) == reqDate
# fnSAPExists
def fnajaxSAPExists(reqDate=date.today()):
    """
    returns true or false to an ajax caller indicating if SAP_SOH data exists for reqDate
    """

    return jsonify(fnSAPExists(reqDate))
# fnajaxSAPExists


@login_required
def fnShowSAP(reqDate=date.today()):

    reqDate = coerce_date(reqDate)
    _myDtFmt = current_app.config.get('DEFAULT_DATEFORMAT', '%Y-%m-%d')

    SAP_tbl = fnSAPList(for_date=reqDate)
    
    stmt = select(SAP_SOHRecs.uploaded_at).distinct().order_by(SAP_SOHRecs.uploaded_at.desc())
    SAPDatesRaw = app_db.session.execute(stmt).scalars().all()
    
    SAPDates = []
    for D in SAPDatesRaw:
        SAPDates.append(D.strftime(_myDtFmt))


    cntext = {'reqDate': SAP_tbl['reqDate'],
            'SAPDateList': SAPDates,
            'SAPDate': SAP_tbl['SAPDate'].strftime(_myDtFmt),
            'SAPSet': SAP_tbl['SAPTable'],
            }
    templt = 'SAP/show_SAP_table.html'
    return checkTemplate_and_render(templt, cntext)


####################################################################################
####################################################################################
####################################################################################


# read the last SAP list before for_date into a list of SAP_SOHRecs
def fnSAPList(for_date = date.today(), matl = None) -> dict:
    """
    finally done!: allow matl to be a MaterialList object or an id
    matl is a Material (string, NOT object!), or list, tuple or queryset of Materials to list, or None if all records are to be listed
    the SAPDate returned is the last one prior or equal to for_date
    """
    _myDtFmt = '%Y-%m-%d %H:%M'

    dateObj = coerce_date(for_date)

    LatestSAPDate = nearestSAPDate(for_date=dateObj)

    # Create the subquery equivalent to Django's Subquery + OuterRef
    uom_subquery = (
        select(UnitsOfMeasure.Multiplier1)
        .where(UnitsOfMeasure.UOM == SAP_SOHRecs.BaseUnitofMeasure)
        .limit(1)
        .scalar_subquery()
    )

    # Build the main query with the annotation (label) and ordering
    stmt = (
        select(
            SAP_SOHRecs, 
            uom_subquery.label("mult")
        )
        .join(SAP_SOHRecs.Material)  # Assuming a relationship is configured
        .where(SAP_SOHRecs.uploaded_at == LatestSAPDate)
        .order_by(
            MaterialList.org_id, 
            MaterialList.Material, 
            SAP_SOHRecs.StorageLocation
        )
    )
    # restrict by material if matl is provided, otherwise get all records
    if matl:
        if isinstance(matl,str):
            raise DeprecationWarning('fnSAPList by Matl string is deprecated')
        elif isinstance(matl,MaterialList):  # handle case matl is a MaterialList instance here
            stmt = stmt.where(SAP_SOHRecs.Material_id == matl.id)
        elif isinstance(matl,int):  # handle case matl is a MaterialList id here
            stmt = stmt.where(SAP_SOHRecs.Material_id == matl)
        else:   # it better be an iterable!
            stmt = stmt.where(SAP_SOHRecs.Material_id.in_([m for m in matl]))
        # endif matl type
    # endif matl provided

    SAPLatest = app_db.session.execute(stmt).all()

    SList = {'reqDate': for_date, 'SAPDate': LatestSAPDate, 'SAPTable':[]}

    # yea, building SList is sorta wasteful, but a lot of existing code depends on it
    # won't be changing it until a total revamp of WICS
    if not SAPLatest:
        SList['SAPDate'] = None
    SList['SAPTable'] = SAPLatest

    return SList


####################################################################################
####################################################################################
####################################################################################

##### the suite of procs to support fnUploadSAP


# class UploadSAPForm(forms.Form):
#     uploaded_at = forms.DateField(widget=forms.widgets.DateInput())
#     SAPFile = forms.FileField()

# def proc_UpSAPSpreadsheet_00InitUpSAP(dbToUse, reqid):
#     acomm = set_async_comm_state(
#         dbToUse,
#         reqid,
#         processname = 'Upload SAP MM52',
#         statecode = 'reading-spreadsht-init',
#         statetext = 'Initializing',
#         new_async=True,
#         )
#     UploadSAPResults.objects.using(dbToUse).all().delete()

# def proc_UpSAPSpreadsheet_00CopyUpSAPSpreadsheet(req, reqid):
#     dbToUse = user_db(req)
#     acomm = set_async_comm_state(
#         dbToUse,
#         reqid,
#         statecode = 'uploading-sprsht',
#         statetext = 'Uploading Spreadsheet',
#         )

#     SAPFile = req.FILES['SAPFile']
#     svdir = getcParm(req, 'SAP-FILELOC')
#     fName = svdir+"tmpSAP"+str(uuid.uuid4())+ExcelWorkbook_fileext
#     with open(fName, "wb") as destination:
#         for chunk in SAPFile.chunks():
#             destination.write(chunk)

#     return fName

# def proc_UpSAPSpreadsheet_01ReadSheet(dbToUse, reqid, fName, UplDate):
#     acomm = set_async_comm_state(
#         dbToUse,
#         reqid,
#         statecode = 'rdng-sprsht',
#         statetext = 'Reading Spreadsheet',
#         )

#     nRowsAdded = 0
#     SprshtRowNum = 0

#     wb = load_workbook(filename=fName, read_only=True)
#     ws = wb.active

#     _SStName_Material = 'Material'
#     _TblName_Material = 'MaterialPartNum'
#     _SStName_Plant = 'Plant'
#     _TblName_Plant = 'Plant'
#     SAPcolmnNames = ws[1]
#     SAPcolmnMap = {_TblName_Material: None, _TblName_Plant: None}
#     SAP_SSName_TableName_map = {
#             _SStName_Material: _TblName_Material,  # Material+org will translate to a Material_id
#             _SStName_Plant: _TblName_Plant,
#             'Material description': 'Description',
#             'Material type': 'MaterialType',
#             'Storage location': 'StorageLocation',
#             'Base Unit of Measure': 'BaseUnitofMeasure',
#             'Unrestricted': 'Amount',
#             'Currency': 'Currency',
#             'Value Unrestricted': 'ValueUnrestricted',
#             'Special Stock': 'SpecialStock',
#             'Blocked':'Blocked',
#             'Value BlockedStock':'ValueBlocked',
#             'Vendor':'Vendor',
#             'Batch': 'Batch',
#             }
#     for col in SAPcolmnNames:
#         if col.value in SAP_SSName_TableName_map:
#             colkey = SAP_SSName_TableName_map[col.value]
#             # has this col.value already been mapped?
#             if (colkey in SAPcolmnMap and SAPcolmnMap[colkey] is not None):
#                 # yes, that's a problem
#                 set_async_comm_state(
#                     dbToUse,
#                     reqid,
#                     statecode = 'fatalerr',
#                     statetext = f'SAP Spreadsheet has bad header row - More than one column named {col.value}.  See Calvin to fix this.',
#                     result = 'FAIL - bad spreadsheet',
#                     )
#                 wb.close()
#                 os.remove(fName)
#                 return
#             else:
#                 SAPcolmnMap[colkey] = col.column - 1
#             # endif previously mapped
#         #endif col.value in SAP_SSName_TableName_map
#     #endfor col in SAPcolmnNames
#     if (SAPcolmnMap[_TblName_Material] is None) or (SAPcolmnMap[_TblName_Plant] is None):   # or SAPcol['StorageLocation'] == None or SAPcol['Amount'] == None):
#         set_async_comm_state(
#             dbToUse,
#             reqid,
#             statecode = 'fatalerr',
#             statetext = f'SAP Spreadsheet has bad header row - no {_SStName_Material} column or no {_SStName_Plant} column.  See Calvin to fix this.',
#             result = 'FAIL - bad spreadsheet',
#             )
#         wb.close()
#         os.remove(fName)
#         return

#     # if SAP SOH records exist for this date, kill them; only one set of SAP SOH records per day
#     # (this was signed off on by user before coming here)
#     SAP_SOHRecs.objects.using(dbToUse).filter(uploaded_at=UplDate).delete()

#     numrows = ws.max_row
#     SprshtRowNum = 1
#     for row in ws.iter_rows(min_row=SprshtRowNum+1, values_only=True):
#         SprshtRowNum += 1
#         if SprshtRowNum % 100 == 0:
#             set_async_comm_state(
#                 dbToUse,
#                 reqid,
#                 statecode = 'rdng-sprsht',
#                 statetext = f'Reading Spreadsheet ... record {SprshtRowNum} of {numrows}<br><progress max="{numrows}" value="{SprshtRowNum}"></progress>',
#                 )

#         if row[SAPcolmnMap[_TblName_Material]]==None: MatlNum = ''
#         else: MatlNum = row[SAPcolmnMap[_TblName_Material]]
#         if len(str(MatlNum)):
#             _org = SAPPlants_org.objects.using(dbToUse).filter(SAPPlant=row[SAPcolmnMap[_TblName_Plant]])[0].org
#             try:
#                 MatlRec = MaterialList.objects.using(dbToUse).get(org=_org,Material=MatlNum)
#             except:
#                 MatlRec = None
#             if not MatlRec:
#                 UploadSAPResults(
#                     errState = 'error',
#                     errmsg = f'either {MatlNum}  does not exist in MaterialList or incorrect Plant ({str(row[SAPcolmnMap[_TblName_Plant]])}) given',
#                     rowNum = SprshtRowNum
#                     ).save(using=dbToUse)
#             else:
#                 SRec = SAP_SOHRecs(
#                         org = _org,     # will be going away - or will it???
#                         uploaded_at = UplDate,
#                         Material = MatlRec
#                         )
#                 for fldName, colNum in SAPcolmnMap.items():
#                     if fldName == _TblName_Material:
#                         pass    # not continue - we are preserving the incoming MaterialPartNum string
#                     if row[colNum] is None: setval = ''
#                     else: setval = row[colNum]
#                     setattr(SRec, fldName, setval)
#                 SRec.save(using=dbToUse)
#                 nRowsAdded += 1
#             # endif MatlRec
#         # endif len(str(MatlNum))
#     #endfor row in ws.iter_rows

#     UploadSAPResults(
#         errState = 'nRowsTotal',
#         errmsg = '',
#         rowNum = SprshtRowNum
#         ).save(using=dbToUse)
#     UploadSAPResults(
#         errState = 'nRowsAdded',
#         errmsg = '',
#         rowNum = nRowsAdded
#         ).save(using=dbToUse)

#     # close and kill temp files
#     wb.close()
#     os.remove(fName)
# def done_UpSAPSpreadsheet_01ReadSheet(t):
#     dbToUse = t.args[0]
#     reqid = t.args[1]
#     statecode = async_comm.objects.using(dbToUse).get(pk=reqid).statecode
#     if statecode != 'fatalerr':
#         set_async_comm_state(
#             dbToUse,
#             reqid,
#             statecode = 'done-rdng-sprsht',
#             statetext = f'Finished Reading Spreadsheet',
#             )
#         proc_UpSAPSpreadsheet_99_FinalProc(dbToUse, reqid)
#     #endif stateocde != 'fatalerr'

# def proc_UpSAPSpreadsheet_99_FinalProc(dbToUse, reqid):
#     set_async_comm_state(
#         dbToUse,
#         reqid,
#         statecode = 'done',
#         statetext = 'Finished Processing Spreadsheet',
#         )

# def proc_UpSAPSpreadsheet_99_Cleanup(dbToUse, reqid):
#     # also kill reqid, acomm, qcluster process
#     async_comm.objects.using(dbToUse).filter(pk=reqid).delete()

#     try:
#         os.kill(int(reqid), signal.SIGTERM)
#     except AttributeError:
#         pass
#     try:
#         os.kill(int(reqid), signal.SIGKILL)
#     except AttributeError:
#         pass

#     # delete the temporary table
#     UploadSAPResults.objects.using(dbToUse).all().delete()

# @login_required
# def fnUploadSAP(req):

#     dbToUse = user_db(req)
#     client_phase = req.POST['phase'] if 'phase' in req.POST else None
#     reqid = req.COOKIES['reqid'] if 'reqid' in req.COOKIES else None
#     UplDate = calvindate(req.POST['uploaded_at']).isoformat() if 'uploaded_at' in req.POST else date.today()

#     if req.method == 'POST':

#         if client_phase=='init-upl':
#             retinfo = HttpResponse()

#             # start django_q broker
#             reqid = subprocess.Popen(
#                 ['python', f'{django_settings.BASE_DIR}/manage.py', 'qcluster']
#             ).pid
#             retinfo.set_cookie('reqid',str(reqid))
#             proc_UpSAPSpreadsheet_00InitUpSAP(dbToUse, reqid)

#             # save the file so we can open it as an excel file
#             fName = proc_UpSAPSpreadsheet_00CopyUpSAPSpreadsheet(req, reqid)

#             task01 = async_task(proc_UpSAPSpreadsheet_01ReadSheet, dbToUse, reqid, fName, UplDate, hook=done_UpSAPSpreadsheet_01ReadSheet)

#             acomm_fake = {
#                 'statecode': 'starting',
#                 'statetext': 'SAP MM60 Update Starting',
#                 }
#             retinfo.write(json.dumps(acomm_fake))
#             return retinfo
#         elif client_phase=='waiting':
#             retinfo = HttpResponse()

#             acomm = async_comm.objects.using(dbToUse).values().get(pk=reqid)    # something's very wrong if this doesn't exist
#             stcode = acomm['statecode']
#             if stcode == 'fatalerr':
#                 pass
#             retinfo.write(json.dumps(acomm))
#             return retinfo
#         elif client_phase=='wantresults':
#             if UploadSAPResults.objects.using(dbToUse).filter(errState = 'nRowsAdded').exists():
#                 nRowsAdded = UploadSAPResults.objects.using(dbToUse).filter(errState = 'nRowsAdded')[0].rowNum
#             else:
#                 nRowsAdded = 0
#             if UploadSAPResults.objects.using(dbToUse).filter(errState = 'nRowsTotal').exists():
#                 nRowsTotal = UploadSAPResults.objects.using(dbToUse).filter(errState = 'nRowsTotal')[0].rowNum
#             else:
#                 nRowsTotal = 0
#             UplResults = UploadSAPResults.objects.using(dbToUse).exclude(errState__in = ['nRowsAdded','nRowsTotal'])
#             cntext = {
#                 'uploaded_at':UplDate,
#                 'nRows':nRowsAdded,
#                 'nRowsRead':nRowsTotal,
#                 'UplProblems': UplResults,
#                     }
#             templt = 'frm_upload_SAP_Success.html'
#             return render(req, templt, cntext)
#         elif client_phase=='resultspresented':
#             proc_UpSAPSpreadsheet_99_Cleanup(dbToUse, reqid)
#             retinfo = HttpResponse()
#             retinfo.delete_cookie('reqid')

#             return retinfo
#         else:
#             return
#         #endif client_phase

#     else:   # req.method != 'POST'
#         LastSAPUpload = SAP_SOHRecs.objects.using(dbToUse).all().aggregate(LastSAPDate=Max('uploaded_at'))
#         # .first().values('LastSAPDate')['LastSAPDate']

#         form = UploadSAPForm()
#         cntext = {'form': form,
#                 'LastSAPUploadDate': LastSAPUpload['LastSAPDate'],
#                 }
#         templt = 'frm_upload_SAP.html'
#     #endif  req.method == 'POST'

#     return render(req, templt, cntext)

